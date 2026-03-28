import openpyxl

from .utils import BaseParser


class Parser(BaseParser):
    """Extract text from Excel files (.xlsx).
    """

    def extract(self, filename, **kwargs):
        workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        output = "\n"
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                new_output = []
                for value in row:
                    if value is not None:
                        new_output.append(str(value))
                if new_output:
                    output += ' '.join(new_output) + '\n'
        workbook.close()
        return output
